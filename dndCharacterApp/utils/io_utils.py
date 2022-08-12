# ----------------------------------------------------------------------------------------------- #
# Module which contains methods for writing/reading/proccessing files such as excel sheets and
# PDF files.
# Also contains a list of file paths used throughout the program.
# ----------------------------------------------------------------------------------------------- #

import openpyxl as pyxl
from os.path import exists
from openpyxl.worksheet.worksheet import Worksheet as worksheet
from PyPDF2 import PdfWriter, PdfFileReader

# List of file paths
FOLDER_EXTENSION = "C:/Users/Public/Documents/DndRandApp-Build-1-Python-8-11-2022/lib/"
RACE_FILE_PATH = "races/Race.xlsx"
TOOL_TYPE_FILE_PATH = "items/ToolTypes.xlsx"
WEAPON_TYPE_FILE_PATH = "items/WeaponTypes.xlsx"
ARMOR_TYPE_FILE_PATH = "items/ArmorTypes.xlsx"
WEAPONS_STATS_FILE_PATH = "items/WeaponStats.xlsx"
ARMOR_STATS_FILE_PATH = "items/ArmorStats.xlsx"
LANGUAGE_FILE_PATH = "general/Languages.xlsx"
BKGR_BONDS_FILE_PATH = "backgrounds/Bonds.xlsx"
BKGR_FLAWS_FILE_PATH = "backgrounds/Flaws.xlsx"
BKGR_IDEALS_FILE_PATH = "backgrounds/Ideals.xlsx"
BKGR_ATRBT_FILE_PATH = "backgrounds/Attributes.xlsx"
BKGR_PERS_TRT_FILE_PATH = "backgrounds/PersonalityTraits.xlsx"
CLASS_STAND_FILE_PATH = "classes/StandardFeatures.xlsx"
CLASS_CHOICE_FILE_PATH = "classes/ChoiceFeatures.xlsx"
CLASS_ATT_SPELL_INCR_FILE_PATH = "classes/AttackAndSpellCastingLevelIncr.xlsx"
CHARACTER_SHEET_FILE_PATH = "general/5E_CharacterSheet_Fillable.pdf"
PDF_STR_FIELD_NAMES = ("ClassLevel", "Background", "CharacterName", "Race ", "Alignment",
			"ProfBonus", "AC", "Initiative", "Speed", "HPMax", "HPCurrent", "HD", "HDTotal", "Passive",
			"PersonalityTraits ", "Ideals", "Bonds", "Flaws",
			"ProficienciesLang", "Equipment",
			"STR", "DEX", "CON", "INT", "WIS", "CHA",
			"STRmod",  "DEXmod ", "CONmod", "INTmod",  "WISmod",  "CHamod",
			"ST Strength", "ST Dexterity", "ST Constitution", "ST Intelligence", "ST Wisdom", "ST Charisma",
			"Acrobatics", "Animal", "Arcana", "Athletics", "Deception ", "History ", "Insight", "Intimidation",
			"Investigation ", "Medicine", "Nature", "Perception ", "Performance", "Persuasion", "Religion",
			"SleightofHand", "Stealth ", "Survival",
			"Wpn Name", "Wpn1 AtkBonus", "Wpn1 Damage", "Wpn Name 2", "Wpn2 AtkBonus ", "Wpn2 Damage ",
			"Wpn Name 3", "Wpn3 AtkBonus  ", "Wpn3 Damage ",
			"CP", "SP", "EP", "GP", "PP", "Features and Traits", "AttacksSpellcasting",
			"Spellcasting Class 2", "SpellcastingAbility 2", "SpellSaveDC  2", "SpellAtkBonus 2")
PDF_BUTTON_FIELD_NAMES = ("Check Box 11", "Check Box 18", "Check Box 19", "Check Box 20", "Check Box 21", "Check Box 22",
				"Check Box 23", "Check Box 24", "Check Box 25", "Check Box 26", "Check Box 27", "Check Box 28",
				"Check Box 29", "Check Box 30", "Check Box 31","Check Box 32", "Check Box 33", "Check Box 34",
				"Check Box 35", "Check Box 36", "Check Box 37", "Check Box 38", "Check Box 39", "Check Box 40")
PDF_STR_SPELL_FIELDS = (
			# Level 0 Spell TextBoxes
			("Spells 1014", "Spells 1016", "Spells 1017", "Spells 1018", "Spells 1019", "Spells 1020", "Spells 1021", "Spells 1022"),
			# Level 1 Spell TextBoxes
			("Spells 1015", "Spells 1023", "Spells 1024", "Spells 1025", "Spells 1026", "Spells 1027", "Spells 1028",
			"Spells 1029", "Spells 1030", "Spells 1031", "Spells 1032", "Spells 1033"),
			# Level 2 Spell TextBoxes
			("Spells 1046", "Spells 1034", "Spells 1035", "Spells 1036", "Spells 1037", "Spells 1038", "Spells 1039",
			"Spells 1040", "Spells 1041", "Spells 1042", "Spells 1043", "Spells 1044", "Spells 1045"),
			# Level 3 Spell TextBoxes
			("Spells 1048", "Spells 1047", "Spells 1049", "Spells 1050", "Spells 1051", "Spells 1052", "Spells 1053", "Spells 1054",
			"Spells 1055", "Spells 1056", "Spells 1057", "Spells 1058", "Spells 1059"),
			# Level 4 Spell TextBoxes
			("Spells 1061", "Spells 1060", "Spells 1062", "Spells 1063", "Spells 1064", "Spells 1065", "Spells 1066", "Spells 1067",
			"Spells 1068", "Spells 1069", "Spells 1070", "Spells 1071", "Spells 1072"),
			# Level 5 Spell TextBoxes
			("Spells 1074", "Spells 1073", "Spells 1075", "Spells 1076", "Spells 1077", "Spells 1078", "Spells 1079", "Spells 1080",
			"Spells 1081"),
			# Level 6 Spell TextBoxes
			("Spells 1083", "Spells 1082", "Spells 1084", "Spells 1085", "Spells 1086", "Spells 1087", "Spells 1088", "Spells 1089",
			"Spells 1090"),
			# Level 7 Spell TextBoxes
			("Spells 1092", "Spells 1091", "Spells 1093", "Spells 1094", "Spells 1095", "Spells 1096", "Spells 1097", "Spells 1098",
			"Spells 1099"),
			# Level 8 Spell TextBoxes
			("Spells 10101", "Spells 10100", "Spells 10102", "Spells 10103", "Spells 10104", "Spells 10105", "Spells 10106"),
			# Level 9 Spell TextBoxes
			("Spells 10108", "Spells 10107", "Spells 10109", "Spells 101010", "Spells 101011", "Spells 101012", "Spells 101013")
			)
PDF_STR_SPELL_SLOT_FIELDS = ("SlotsTotal 19", "SlotsTotal 20", "SlotsTotal 21",
			"SlotsTotal 22", "SlotsTotal 23", "SlotsTotal 24", "SlotsTotal 25", "SlotsTotal 26", "SlotsTotal 27")
PDF_BUTTON_SPELL_FIELDS = (
			# Level 1 Prepared Spell Buttons
			("Check Box 251", "Check Box 309", "Check Box 3010", "Check Box 3011", "Check Box 3012",
			"Check Box 3013", "Check Box 3014", "Check Box 3015", "Check Box 3016", "Check Box 3017",
			"Check Box 3018", "Check Box 3019"),
			# Level 2 Prepared Spell Buttons
			("Check Box 313", "Check Box 310", "Check Box 3020", "Check Box 3021", "Check Box 3022",
			"Check Box 3023", "Check Box 3024", "Check Box 3025", "Check Box 3026", "Check Box 3027",
			"Check Box 3028", "Check Box 3029", "Check Box 3030"),
			# Level 3 Prepared Spell Buttons
			("Check Box 315", "Check Box 314", "Check Box 3031", "Check Box 3032", "Check Box 3033", "Check Box 3034", "Check Box 3035",
			"Check Box 3036", "Check Box 3037", "Check Box 3038", "Check Box 3039", "Check Box 3040", "Check Box 3041"),
			# Level 4 Prepared Spell Buttons
			("Check Box 317", "Check Box 316", "Check Box 3042", "Check Box 3043", "Check Box 3044", "Check Box 3045", "Check Box 3046",
			"Check Box 3047", "Check Box 3048", "Check Box 3049", "Check Box 3050", "Check Box 3051", "Check Box 3052"),
			# Level 5 Prepared Spell Buttons
			("Check Box 319", "Check Box 318", "Check Box 3053", "Check Box 3054", "Check Box 3055", "Check Box 3056", "Check Box 3057",
			"Check Box 3058", "Check Box 3059"),
			# Level 6 Prepared Spell Buttons
			("Check Box 321", "Check Box 320", "Check Box 3060", "Check Box 3061", "Check Box 3062", "Check Box 3063", "Check Box 3064",
			"Check Box 3065", "Check Box 3066"),
			# Level 7 Prepared Spell Buttons
			("Check Box 323", "Check Box 322", "Check Box 3067", "Check Box 3068", "Check Box 3069", "Check Box 3070", "Check Box 3071",
			"Check Box 3072", "Check Box 3073"),
			# Level 8 Prepared Spell Buttons
			("Check Box 325", "Check Box 324", "Check Box 3074", "Check Box 3075", "Check Box 3076", "Check Box 3077", "Check Box 3078"),
			# Level 9 Prepared Spell Buttons
			("Check Box 327", "Check Box 326", "Check Box 3079", "Check Box 3080", "Check Box 3081", "Check Box 3082", "Check Box 3083")
			)
PDF_FILE_PATH = "general/5E_CharacterSheet_Fillable.pdf"


# Function gets the contents of a column given either the name of the first element
# in that column or an index.
# If a string is given as an argument then a binary/linear search is preformed whether
# the file is sorted or not to find its index
def get_col(elementOrIndex, file_path: str, isSorted: bool) -> ():
    # Absolute path is need to find the location of the resource folder

    # If program is running from IDE use folder in project folder
    if (exists("../../lib/" + file_path)):
        absolute_path = "../../lib" + file_path
    # Use external folder
    else:
        absolute_path = FOLDER_EXTENSION + file_path

    sheet = _get_excel_sheet(absolute_path)

    if sheet is not None:
        # Get the index if the argument is a String
        if isinstance(elementOrIndex, str):
            # Row is retrieved to search through first entries of all columns
            row_Contents = _get_row_contents(sheet, 1)
            # Get index
            elementOrIndex = \
                _get_index_binary_search(elementOrIndex, row_Contents) if isSorted \
                    else _get_index(elementOrIndex, row_Contents)

        if elementOrIndex != -1:
            return _get_column_contents(sheet, elementOrIndex)

    return None


# Function gets the contents of a row given an index and a file name.
def get_row(index: int, file_path: str):
    # Absolute path is need to find the location of the resource folder
    absolute_path = FOLDER_EXTENSION + file_path
    sheet = _get_excel_sheet(absolute_path)

    if sheet is not None:
        return _get_row_contents(sheet, index)


# Function used to get an Excel sheet from a given file.
# None is returned if any exception occurs.
def _get_excel_sheet(file_path: str) -> worksheet:
    # First try to open the workbook if successful,
    # Return the sheet and then close the workbook after.
    try:
        wb = pyxl.load_workbook(file_path)
        sheet = wb.active
        wb.close()
        return sheet
    except FileNotFoundError:
        return None
    except IOError:
        return None


# Function used to get the contents of a specific row given a worksheet and
# index. Indexing begins at 1 instead of 0.
# Returns a tuple containing the contents of said row.
def _get_row_contents(sheet: worksheet, index: int) -> ():
    # openpyxl has rows and columns reversed for some reason.
    max_row = sheet.max_column
    row = []

    if index > max_row:
        return None

    for i in range(index, max_row + 1):
        cell = sheet.cell(row=index, column=i)
        # Do not add to the list if the cell is empty.
        if cell is not None:
            # Convert all numeric values to an integer
            if isinstance(cell.value, float):
                row.append(int(cell.value))
            else:
                row.append(cell.value)

    return tuple(row)


# Function used to get the contents of a specific column given a worksheet and
# index. Indexing begins at 1 instead of 0.
# Returns a tuple containing the contents of said column.
def _get_column_contents(sheet: worksheet, index: int) -> ():
    # openpyxl has rows and columns reversed for some reason.
    max_column = sheet.max_row
    column = []

    for i in range(1, max_column + 1):
        cell = sheet.cell(row=i, column=index)
        # Do not add to the list if the cell is empty.
        if cell is not None and cell.value is not None:
            # Convert all numeric values to an integer
            if isinstance(cell.value, float):
                column.append(int(cell.value))
            else:
                column.append(cell.value)

    return tuple(column)


# Function used to linearly search for an index in a tuple for a given element
# -1 is returned if the element is not in the tuple
def _get_index(element: str, contents: ()) -> int:
    for i in range(len(contents)):
        if str(contents[i]).lower() == element.lower():
            return i + 1
    return -1


# Function used preform a binary search for an index in a tuple for a given element
# -1 is returned if the element is not in the tuple
def _get_index_binary_search(element: str, contents: ()) -> int:
    low = 0
    high = len(contents) - 1

    while low <= high:
        guess_index = int(low + ((high - low) / 2))
        guess = contents[guess_index]

        if str(guess).lower() < element.lower():
            low = guess_index + 1
        elif str(guess).lower() > element.lower():
            high = guess_index - 1
        else:
            return guess_index + 1

    return -1

# Function used to fill pdf with information gathered from a complete character sheet.
# A boolean is returned to notify the sender of the success.
def create_character_sheet(string_field_contents: (), button_field_contents:(), spells: (),
                           spell_slots: (), prepared_known_caster: str, out_file: str) -> bool:
    # Absolute path of file is specified
    absolute_path = FOLDER_EXTENSION + PDF_FILE_PATH

    # If the pdf can be opened continue writing to it, if not return false.
    try:
        pdf = PdfFileReader(open(absolute_path, 'rb'))
    except (PermissionError, FileNotFoundError):
        pdf = None

    if pdf is not None:
        # Create a new pdf to be written to using the base provided,
        # add all the necessary pages, in this case two.
        wr = PdfWriter()
        for i in range(2):
            page = pdf.pages[i]
            wr.addPage(page)

        # Set proper page number which fields will be accessed by
        page_num = 0

        # Write the string values to the fields mostly on page 1.
        for i in range(len(PDF_STR_FIELD_NAMES)):
            # Last section of field names is on second page
            if PDF_STR_FIELD_NAMES[i] == "Spellcasting Class 2":
                page_num = 1

            wr.update_page_form_field_values(
                wr.pages[page_num], {PDF_STR_FIELD_NAMES[i]: string_field_contents[i]}
            )
        page_num = 0
        # Set radio buttons to proper settings on the first page
        for i in range(len(PDF_BUTTON_FIELD_NAMES)):
            # Value for a selected radio button is "/Yes"
            if button_field_contents[i]:
                value = "/Yes"
            else:
                value = ""
            wr.update_page_form_field_values(
                wr.pages[page_num], {PDF_BUTTON_FIELD_NAMES[i]: value}
            )

        page_num = 1

        # Fill the character's spells if they are a caster.
        if spells is not None and spell_slots is not None:
            # Get max level spell that the character can cast
            spell_level_length = len(spells)

            for i in range(len(PDF_STR_SPELL_FIELDS)):
                # Stop writing if max spell has been exceeded.
                if i >= spell_level_length:
                    break

                # Get length of the list of spells the character can cast for that
                # spell level
                len_spells = len(spells[i])

                for j in range(len(PDF_STR_SPELL_FIELDS[i])):
                    # Stop writing if the list has ended.
                    if j >= len_spells:
                        break
                    wr.update_page_form_field_values(
                        wr.pages[page_num], {PDF_STR_SPELL_FIELDS[i][j]: spells[i][j]}
                    )
                    # Check the radio buttons for prepared casters
                    if i != 0 and prepared_known_caster.lower() == "p":
                        wr.update_page_form_field_values(
                            wr.pages[page_num], {PDF_BUTTON_SPELL_FIELDS[i - 1][j]: "/Yes"}
                        )
            # Get the max number of spells the caster has access to.
            slot_length = len(spell_slots)
            for i in range(len(PDF_STR_SPELL_SLOT_FIELDS)):
                # Stop writing if max spell has been exceeded.
                # Cantrips (0-level) spells do not have slots so skip the first entry.
                if i + 1 >= slot_length:
                    break
                wr.update_page_form_field_values(
                    wr.pages[page_num], {PDF_STR_SPELL_SLOT_FIELDS[i]: str(spell_slots[i + 1])}
                )

        # Write the contents to the given destination
        if out_file is not None:
            with open(out_file, "wb") as output_stream:
                wr.write(output_stream)
        else:
            return False

        return True
    else:
        return False
